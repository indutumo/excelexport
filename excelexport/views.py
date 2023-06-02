from django.shortcuts import render
from django.http.response import HttpResponse
from .models import *
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
from openpyxl.styles import *
import decimal


def is_valid_queryparam(param):
    return param != '' and param is not None


def countries_gdp_list(request):
    qs = CountryGDP.objects.order_by('name')

    name = request.GET.get('name')
    year = request.GET.get('year')

    request.session['name'] = name
    request.session['year'] = year

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 30)

    try:
        qs = paginator.page(page)
    except PageNotAnInteger:
        qs = paginator.page(1)
    except EmptyPage:
        qs = paginator.page(paginator.num_pages)

    context = {
        'countries_list': qs,
        'name': name,
        'year':year,
    }
    return render(request, "excelexport/countries_list.html", context)




def countries_gdp_excel(request):
    qs = CountryGDP.objects.order_by('name')

    if 'name' in request.session:
        name = request.session['name']
    else:
        name = None

    if 'year' in request.session:
        year = request.session['year']
    else:
        year = None

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    if year is None or year == '':
        year = "2013 - 2016"
    else:
        year = year

    if name is None or name == '':
        name = "All Countries"
    else:
        name = name

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'Countries GDP List' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')
    first_cell = worksheet['A1']
    first_cell.value = "Countries GDP List" + " " + year
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value = name
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Countries GDP List' + " " + year

    # Define the titles for columns
    columns = ['Country Name','Country Code','Year', 'Value in USD']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['D3']
        third_cell.alignment = Alignment(horizontal="right")

    for countries in qs:
        row_num += 1

        # Define the data for each cell in the row
        row = [countries.name,countries.code,countries.year,countries.value]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
        	cell = worksheet.cell(row=row_num, column=col_num)
        	cell.value = cell_value
	        if isinstance(cell_value, decimal.Decimal):
	            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    workbook.save(response)
    return response
